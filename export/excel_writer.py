from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from export.excel_schema import APPLICATION_COLUMNS, EMAIL_LOG_COLUMNS


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _to_excel_dt(value: Any) -> Any:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value).strip()
    for candidate in (text.replace("Z", "+00:00"), text):
        try:
            dt = datetime.fromisoformat(candidate)
            return dt.replace(tzinfo=None)
        except ValueError:
            continue
    return text


def read_existing_application_overrides(path: str) -> dict[str, dict[str, Any]]:
    out_path = Path(path)
    if not out_path.exists():
        return {}
    wb = load_workbook(out_path)
    if "Applications" not in wb.sheetnames:
        return {}
    ws = wb["Applications"]
    header = [c.value for c in ws[1]]
    idx = {str(name): i for i, name in enumerate(header) if name}
    if "RecordID" not in idx:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = row[idx["RecordID"]]
        if not rid:
            continue
        out[str(rid)] = {
            "Status": row[idx["Status"]] if "Status" in idx else "",
            "Notes": row[idx["Notes"]] if "Notes" in idx else "",
            "NextStep": row[idx["NextStep"]] if "NextStep" in idx else "",
            "UserLockStatus": int(row[idx["UserLockStatus"]] or 0) if "UserLockStatus" in idx else 0,
            "UserLockNotes": int(row[idx["UserLockNotes"]] or 0) if "UserLockNotes" in idx else 0,
            "UserLockNextStep": int(row[idx["UserLockNextStep"]] or 0) if "UserLockNextStep" in idx else 0,
        }
    return out


def _build_application_row(row: Any, last_detected_type: str, tracking_category: str) -> list[Any]:
    return [
        row["record_id"],
        row["company"],
        row["role"],
        row["location"],
        row["req_id"],
        row["job_url"],
        row["source"],
        row["status"],
        _to_excel_dt(row["status_date"]),
        _to_excel_dt(row["date_first_seen"]),
        _to_excel_dt(row["date_applied"]),
        _to_excel_dt(row["last_email_date"]),
        _to_excel_dt(row["follow_up_due"]),
        "",  # DaysSinceLastEmail formula populated post-append.
        row["recruiter_name"] if "recruiter_name" in row.keys() else "",
        row["recruiter_email"] if "recruiter_email" in row.keys() else "",
        row["next_step"],
        row["notes"],
        row["email_thread_link"],
        row["last_message_id_header"],
        row["confidence"],
        row["matched_by"],
        last_detected_type,
        tracking_category,
        row["user_lock_status"],
        row["user_lock_notes"],
        row["user_lock_next_step"],
    ]


def _style_header(ws) -> None:
    ws.row_dimensions[1].height = 24
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT


def _apply_column_widths(ws, headers: list[str]) -> None:
    narrow = {"DateFirstSeen", "DateApplied", "StatusDate", "LastEmailDate", "FollowUpDue", "Confidence", "DaysSinceLastEmail"}
    medium = {"Company", "Status", "Source", "Location", "RecruiterName", "RecruiterEmail", "ReqID_JobID"}
    wide = {"Role", "NextStep", "Notes", "JobURL", "EmailThreadLink", "RawSnippet", "Subject"}
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        if h in narrow:
            ws.column_dimensions[col].width = 13
        elif h in medium:
            ws.column_dimensions[col].width = 18
        elif h in wide:
            ws.column_dimensions[col].width = 44 if h in {"Notes", "RawSnippet"} else 34
        else:
            ws.column_dimensions[col].width = 20


def _add_table(ws, table_name: str) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{max(2, ws.max_row)}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _add_status_validation_and_rules(ws, headers: list[str]) -> None:
    status_idx = headers.index("Status") + 1
    status_col = get_column_letter(status_idx)
    status_validation = DataValidation(
        type="list",
        formula1='"Opportunity,Applied,Interview,Offer,Rejected,Closed,NeedsReview"',
        allow_blank=False,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f"{status_col}2:{status_col}{max(2, ws.max_row)}")

    full_range = f"A2:{get_column_letter(ws.max_column)}{max(2, ws.max_row)}"
    rules = [
        ('=$' + status_col + '2="Interview"', PatternFill("solid", fgColor="C6EFCE")),
        ('=$' + status_col + '2="Offer"', PatternFill("solid", fgColor="C6EFCE")),
        ('=$' + status_col + '2="Rejected"', PatternFill("solid", fgColor="FFC7CE")),
        ('=$' + status_col + '2="NeedsReview"', PatternFill("solid", fgColor="FFEB9C")),
        ('=$' + status_col + '2="Closed"', PatternFill("solid", fgColor="D9D9D9")),
        ('=$' + status_col + '2="Applied"', PatternFill("solid", fgColor="DDEBF7")),
    ]
    for formula, fill in rules:
        ws.conditional_formatting.add(full_range, FormulaRule(formula=[formula], fill=fill))


def _add_followup_rules(ws, headers: list[str]) -> None:
    col = get_column_letter(headers.index("FollowUpDue") + 1)
    full_range = f"A2:{get_column_letter(ws.max_column)}{max(2, ws.max_row)}"
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(
            formula=[f'=AND(${col}2<>"",${col}2<TODAY())'],
            fill=PatternFill("solid", fgColor="FFC7CE"),
        ),
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(
            formula=[f'=AND(${col}2<>"",${col}2>=TODAY(),${col}2<=TODAY()+3)'],
            fill=PatternFill("solid", fgColor="FFEB9C"),
        ),
    )


def _add_days_since_rules(ws, headers: list[str]) -> None:
    col = get_column_letter(headers.index("DaysSinceLastEmail") + 1)
    full_range = f"A2:{get_column_letter(ws.max_column)}{max(2, ws.max_row)}"
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'=AND(${col}2<>"",${col}2>=14)'], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'=AND(${col}2<>"",${col}2>=7,${col}2<14)'], fill=PatternFill("solid", fgColor="FCE4D6")),
    )


def _apply_wrap_for_columns(ws, headers: list[str], col_names: set[str]) -> None:
    for name in col_names:
        if name not in headers:
            continue
        idx = headers.index(name) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(r, idx).alignment = Alignment(wrap_text=True, vertical="top")


def _group_and_hide_audit_columns(ws, headers: list[str]) -> None:
    audit_cols = [
        "EmailThreadLink",
        "LastMessageID",
        "Confidence",
        "MatchedBy",
        "LastDetectedType",
        "TrackingCategory",
        "UserLockStatus",
        "UserLockNotes",
        "UserLockNextStep",
    ]
    for name in audit_cols:
        if name not in headers:
            continue
        idx = headers.index(name) + 1
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].outlineLevel = 1
        ws.column_dimensions[letter].hidden = True
    ws.sheet_properties.outlinePr.summaryRight = True


def _add_dashboard_sheet(wb: Workbook, applications: list[Any]) -> None:
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Job Tracker Dashboard"
    ws["A1"].font = Font(size=16, bold=True)

    status_counter = Counter((r["status"] or "Unknown") for r in applications)
    source_counter = Counter((r["source"] or "Unknown") for r in applications)

    metrics = [
        ("Total Records", len(applications)),
        ("Applied", status_counter.get("Applied", 0)),
        ("Interview", status_counter.get("Interview", 0)),
        ("Offer", status_counter.get("Offer", 0)),
        ("Rejected", status_counter.get("Rejected", 0)),
        ("NeedsReview", status_counter.get("NeedsReview", 0)),
    ]
    ws["A3"] = "Metrics"
    ws["A3"].font = Font(bold=True)
    row = 4
    for label, val in metrics:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = val
        row += 1

    ws["A11"] = "Status Counts"
    ws["A11"].font = Font(bold=True)
    r = 12
    for k, v in status_counter.most_common():
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        r += 1

    ws["D11"] = "Applications by Source"
    ws["D11"].font = Font(bold=True)
    r2 = 12
    for k, v in source_counter.most_common(12):
        ws[f"D{r2}"] = k
        ws[f"E{r2}"] = v
        r2 += 1

    if r > 12:
        chart1 = BarChart()
        chart1.title = "Status Counts"
        chart1.y_axis.title = "Count"
        data = Reference(ws, min_col=2, min_row=11, max_row=r - 1)
        cats = Reference(ws, min_col=1, min_row=12, max_row=r - 1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 6
        chart1.width = 10
        ws.add_chart(chart1, "G3")

    if r2 > 12:
        chart2 = BarChart()
        chart2.title = "Applications by Source"
        chart2.y_axis.title = "Count"
        data2 = Reference(ws, min_col=5, min_row=11, max_row=r2 - 1)
        cats2 = Reference(ws, min_col=4, min_row=12, max_row=r2 - 1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.height = 6
        chart2.width = 10
        ws.add_chart(chart2, "G18")

    # Visual legend for status/follow-up conditional colors.
    ws["A20"] = "Color Legend"
    ws["A20"].font = Font(bold=True)
    legends = [
        ("Interview / Offer", "C6EFCE"),
        ("Applied", "DDEBF7"),
        ("Rejected", "FFC7CE"),
        ("NeedsReview", "FFEB9C"),
        ("Closed", "D9D9D9"),
        ("FollowUpDue <= 3 days", "FFEB9C"),
        ("FollowUpDue overdue", "FFC7CE"),
        ("DaysSinceLastEmail >= 7", "FCE4D6"),
        ("DaysSinceLastEmail >= 14", "FFC7CE"),
    ]
    base = 21
    for i, (label, color) in enumerate(legends):
        r_legend = base + i
        ws[f"A{r_legend}"] = label
        ws[f"B{r_legend}"] = ""
        ws[f"B{r_legend}"].fill = PatternFill("solid", fgColor=color)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 14


def write_excel(
    path: str,
    applications: list[Any],
    email_events: list[Any],
    email_log_limit: int = 5000,
) -> None:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    app_headers = APPLICATION_COLUMNS
    email_headers = EMAIL_LOG_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(app_headers)
    _style_header(ws)
    ws.freeze_panes = "B2"

    event_type_by_uid: dict[int, str] = {}
    event_type_by_message: dict[str, str] = {}
    for ev in email_events:
        event_type_by_uid[int(ev["uid"])] = ev["detected_type"]
        if ev["message_id_header"]:
            event_type_by_message[str(ev["message_id_header"])] = ev["detected_type"]

    col_idx = {h: i + 1 for i, h in enumerate(app_headers)}
    last_email_col = col_idx["LastEmailDate"]
    days_col = col_idx["DaysSinceLastEmail"]
    for row in applications:
        last_type = event_type_by_uid.get(int(row["last_uid"] or 0), event_type_by_message.get(row["last_message_id_header"] or "", "Other"))
        if last_type == "Opportunity":
            tracking_category = "Opportunity Notification"
        elif last_type == "ApplicationConfirmation":
            tracking_category = "Application Confirmation"
        elif last_type in {"InterviewRequest", "Rejection", "Offer"}:
            tracking_category = "Application Update"
        else:
            tracking_category = "Needs Review"

        ws.append(_build_application_row(row, last_type, tracking_category))
        row_idx = ws.max_row

        # Formula: days since last email.
        last_col_letter = get_column_letter(last_email_col)
        ws.cell(row_idx, days_col).value = f'=IF({last_col_letter}{row_idx}="","",TODAY()-{last_col_letter}{row_idx})'

        if row["job_url"]:
            ws.cell(row_idx, col_idx["JobURL"]).hyperlink = row["job_url"]
            ws.cell(row_idx, col_idx["JobURL"]).style = "Hyperlink"
        if row["email_thread_link"]:
            ws.cell(row_idx, col_idx["EmailThreadLink"]).hyperlink = row["email_thread_link"]
            ws.cell(row_idx, col_idx["EmailThreadLink"]).style = "Hyperlink"

    # Date/number formats — single pass over all rows.
    date_col_indices = [col_idx[c] for c in ("StatusDate", "DateFirstSeen", "DateApplied", "LastEmailDate", "FollowUpDue")]
    conf_idx = col_idx["Confidence"]
    days_idx = col_idx["DaysSinceLastEmail"]
    for r in range(2, ws.max_row + 1):
        for dc in date_col_indices:
            ws.cell(r, dc).number_format = "yyyy-mm-dd"
        ws.cell(r, conf_idx).number_format = "0.00"
        ws.cell(r, days_idx).number_format = "0"

    _add_status_validation_and_rules(ws, app_headers)
    _add_followup_rules(ws, app_headers)
    _add_days_since_rules(ws, app_headers)
    _apply_wrap_for_columns(ws, app_headers, {"Notes", "NextStep"})
    _apply_column_widths(ws, app_headers)
    _group_and_hide_audit_columns(ws, app_headers)
    _add_table(ws, "ApplicationsTable")

    log_ws = wb.create_sheet("EmailLog")
    log_ws.append(email_headers)
    _style_header(log_ws)
    log_ws.freeze_panes = "A2"

    count = 0
    for ev in email_events:
        if count >= email_log_limit:
            break
        if ev["detected_type"] == "Other":
            continue
        row = {
            "MessageID": ev["message_id_header"],
            "ThreadID": ev["thread_hint"],
            "ReceivedDate": _to_excel_dt(ev["internal_date"]),
            "From": ev["from_email"],
            "Subject": ev["subject"],
            "DetectedType": ev["detected_type"],
            "LinkedRecordID": ev["linked_record_id"],
            "ExtractorNotes": ev["matched_by"],
            "RawSnippet": ev["snippet"],
        }
        log_ws.append([row.get(h, "") for h in email_headers])
        count += 1
    if "ReceivedDate" in email_headers:
        idx = email_headers.index("ReceivedDate") + 1
        for r in range(2, log_ws.max_row + 1):
            log_ws.cell(r, idx).number_format = "yyyy-mm-dd hh:mm"

    _apply_wrap_for_columns(log_ws, email_headers, {"Subject", "RawSnippet"})
    _apply_column_widths(log_ws, email_headers)
    _add_table(log_ws, "EmailLogTable")

    _add_dashboard_sheet(wb, applications)
    wb.save(path)
